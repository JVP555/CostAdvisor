"""Payload-agnostic Excel I/O for the sheet round-trip mechanism (Scrum 27b).

Styling conventions (fonts/fills/borders) copied from
routers/suppliers.py's export-excel — the only other place in this repo
that writes a styled .xlsx — so this doesn't invent a second visual
language. Cell/sheet protection is new to the codebase but uses openpyxl's
already-installed API.
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Protection
from openpyxl.utils import get_column_letter

from app.services.file_parser import _read_file
from app.services.sheet_roundtrip.base import SheetPayloadSpec

# Column-kind fills — key = dark (matches suppliers.py's header_fill), editable =
# teal (matches suppliers.py's gap_neg_font accent, repurposed as "this is yours
# to edit"), readonly = light grey (context, not an accent color at all).
_FILL_KEY = PatternFill(start_color="2D3436", end_color="2D3436", fill_type="solid")
_FILL_EDITABLE = PatternFill(start_color="00B894", end_color="00B894", fill_type="solid")
_FILL_READONLY = PatternFill(start_color="DFE6E9", end_color="DFE6E9", fill_type="solid")
_FONT_ON_DARK = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_FONT_ON_LIGHT = Font(name="Calibri", bold=True, size=11, color="2D3436")
_FONT_READONLY_DATA = Font(name="Calibri", size=11, italic=True, color="636E72")
_FONT_DATA = Font(name="Calibri", size=11)
_CENTER = Alignment(horizontal="center", vertical="center")


def _normalize_header(label: str) -> str:
    """Same normalization _read_file applies to uploaded headers, so an
    exported header_label round-trips back to the right column regardless
    of the exact wording chosen for display."""
    return label.strip().lower().replace(" ", "_")


def build_export_workbook(spec: SheetPayloadSpec, rows: list[dict]) -> io.BytesIO:
    """Build a .xlsx for the given rows. Key and readonly columns are
    genuinely locked (Excel-level cell protection); only editable columns
    are unlocked. This is a UX guard against fat-fingering, NOT a security
    boundary — sheet protection has no password and is trivially removed in
    Excel. The import/diff step is the real authority: a readonly-column
    edit that survives protection removal is still caught and rejected
    there (see diff.py), never silently applied.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = spec.sheet_name[:31]

    for col_idx, col in enumerate(spec.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=f"{col.header_label} ({col.kind})")
        cell.font = _FONT_ON_DARK if col.kind == "key" else (
            _FONT_ON_LIGHT if col.kind == "editable" else _FONT_ON_LIGHT
        )
        cell.fill = _FILL_KEY if col.kind == "key" else (
            _FILL_EDITABLE if col.kind == "editable" else _FILL_READONLY
        )
        cell.alignment = _CENTER
        cell.protection = Protection(locked=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(col.header_label) + 4)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col in enumerate(spec.columns, start=1):
            value = row.get(col.name)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = _FONT_DATA if col.kind != "readonly" else _FONT_READONLY_DATA
            cell.protection = Protection(locked=(col.kind != "editable"))

    ws.freeze_panes = "A2"
    ws.protection.sheet = True

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def read_import_rows(content: bytes, filename: str, spec: SheetPayloadSpec) -> list[dict]:
    """Read a reimported sheet back into row dicts keyed by internal column
    name (not header_label — matching is by normalized header text, so the
    display wording of a header is decoupled from the column's identity).

    Deliberately does NOT scan for a header-row offset: we control the
    export format and always put headers on row 1, so unlike a tool reading
    a foreign spreadsheet, there's no ambiguity to resolve here. Extra/
    unrecognized columns in the reimported file are silently ignored (e.g.
    Excel's own row-number gutter); missing key columns are a structural
    failure, not a per-row one.
    """
    df = _read_file(content, filename)

    column_to_header = {c: _normalize_header(f"{c.header_label} ({c.kind})") for c in spec.columns}
    header_to_column = {header: c for c, header in column_to_header.items()}
    found = {header: col for header, col in header_to_column.items() if header in df.columns}

    missing_keys = [c.name for c in spec.key_columns if column_to_header[c] not in df.columns]
    if missing_keys:
        raise ValueError(f"Missing required key column(s) in uploaded sheet: {', '.join(missing_keys)}")

    rows = []
    for _, series in df.iterrows():
        row = {}
        for header, col in found.items():
            raw = series[header]
            row[col.name] = None if _is_blank(raw) else raw
        rows.append(row)
    return rows


def _is_blank(value) -> bool:
    if value is None:
        return True
    try:
        import math
        return isinstance(value, float) and math.isnan(value)
    except TypeError:
        return False
