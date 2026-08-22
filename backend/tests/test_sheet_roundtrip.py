"""Scrum 27b — sheet round-trip mechanism (export -> edit offline -> reimport
-> diff -> apply), exercised against the one registered payload
(FormulaRegionCoverage base-price editing). One or more tests per acceptance
criterion, plus the concurrency behavior the ticket calls out explicitly.
"""
from __future__ import annotations

import io
import uuid

import openpyxl
import pytest

from app.models.chemical_family import ChemicalFamily
from app.models.subfamily import Subfamily
from app.models.formula_template import FormulaTemplate, FormulaRegionCoverage
from app.models.sheet_import_run import SheetImportRun

XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture
def coverage_setup(db, tenant_a):
    suffix = uuid.uuid4().hex[:8]
    family = ChemicalFamily(name=f"Fam-{suffix}")
    db.add(family)
    db.flush()
    sub_a = Subfamily(family_id=family.id, name=f"SubA-{suffix}")
    sub_b = Subfamily(family_id=family.id, name=f"SubB-{suffix}")
    db.add_all([sub_a, sub_b])
    db.flush()

    t1 = FormulaTemplate(team_id=None, created_by=tenant_a["user_id"], name=f"T1-{suffix}",
                          code=f"T1-{suffix}", subfamily_id=sub_a.id)
    t2 = FormulaTemplate(team_id=None, created_by=tenant_a["user_id"], name=f"T2-{suffix}",
                          code=f"T2-{suffix}", subfamily_id=sub_a.id)
    t3 = FormulaTemplate(team_id=None, created_by=tenant_a["user_id"], name=f"T3-{suffix}",
                          code=f"T3-{suffix}", subfamily_id=sub_b.id)
    db.add_all([t1, t2, t3])
    db.flush()

    c1 = FormulaRegionCoverage(template_id=t1.id, region="Europe", base_price=100, currency="USD",
                                margin_pct=10, base_year=2024, base_quarter=1,
                                needs_review=False, data_confidence="CONF-HIGH")
    c2 = FormulaRegionCoverage(template_id=t2.id, region="Europe", base_price=200, currency="USD",
                                margin_pct=12, base_year=2024, base_quarter=1,
                                needs_review=True, data_confidence="CONF-LOW")
    c3 = FormulaRegionCoverage(template_id=t3.id, region="Europe", base_price=300, currency="USD",
                                margin_pct=8, base_year=2024, base_quarter=1,
                                needs_review=False, data_confidence="CONF-HIGH")
    db.add_all([c1, c2, c3])
    db.commit()

    yield {"family": family, "sub_a": sub_a, "sub_b": sub_b, "t1": t1, "t2": t2, "t3": t3}

    for m in (FormulaRegionCoverage,):
        db.query(m).filter(m.template_id.in_([t1.id, t2.id, t3.id])).delete(synchronize_session=False)
    db.query(FormulaTemplate).filter(FormulaTemplate.id.in_([t1.id, t2.id, t3.id])).delete(synchronize_session=False)
    db.query(Subfamily).filter(Subfamily.id.in_([sub_a.id, sub_b.id])).delete(synchronize_session=False)
    db.query(ChemicalFamily).filter(ChemicalFamily.id == family.id).delete(synchronize_session=False)
    db.commit()


@pytest.fixture
def admin(user_factory, db):
    """A super-admin user, with any SheetImportRun rows it creates cleaned up
    before user_factory's own teardown deletes the user row (sheet_import_runs
    has no CASCADE on imported_by/applied_by — it's platform audit-trail data
    that should outlive a deleted user in production, so the FK is plain)."""
    u = user_factory(is_super_admin=True)
    yield u
    db.query(SheetImportRun).filter(
        (SheetImportRun.imported_by == u["user_id"]) | (SheetImportRun.applied_by == u["user_id"])
    ).delete(synchronize_session=False)
    db.commit()


def _export(client, subfamily_id=None, needs_review=None):
    params = {}
    if subfamily_id is not None:
        params["subfamily_id"] = subfamily_id
    if needs_review is not None:
        params["needs_review"] = needs_review
    r = client.get("/api/sheets/formula_coverage_price/export", params=params)
    assert r.status_code == 200, r.text
    return r.content


def _load(content: bytes):
    return openpyxl.load_workbook(io.BytesIO(content))


def _find_col(ws, label_prefix: str) -> int:
    for cell in ws[1]:
        if cell.value and str(cell.value).startswith(label_prefix):
            return cell.column
    raise AssertionError(f"column '{label_prefix}' not found in header row")


def _save(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _import(client, content: bytes, subfamily_id=None, needs_review=None):
    params = {}
    if subfamily_id is not None:
        params["subfamily_id"] = subfamily_id
    if needs_review is not None:
        params["needs_review"] = needs_review
    r = client.post(
        "/api/sheets/formula_coverage_price/import",
        params=params,
        files={"file": ("edited.xlsx", content, XLSX_MEDIA)},
    )
    return r


# ── AC1 ──────────────────────────────────────────────────────────────────

def test_export_filters_by_subfamily_and_needs_review(client_as, admin, coverage_setup):
    c = client_as(admin)
    content = _export(c, subfamily_id=coverage_setup["sub_a"].id)
    wb = _load(content)
    ws = wb.active
    code_col = _find_col(ws, "Formula Code")
    codes = {row[code_col - 1].value for row in ws.iter_rows(min_row=2)}
    assert codes == {coverage_setup["t1"].code, coverage_setup["t2"].code}

    content2 = _export(c, subfamily_id=coverage_setup["sub_a"].id, needs_review=True)
    wb2 = _load(content2)
    ws2 = wb2.active
    codes2 = {row[code_col - 1].value for row in ws2.iter_rows(min_row=2)}
    assert codes2 == {coverage_setup["t2"].code}


def test_export_locks_readonly_and_key_columns_not_editable_columns(client_as, admin, coverage_setup):
    c = client_as(admin)
    content = _export(c, subfamily_id=coverage_setup["sub_a"].id)
    wb = _load(content)
    ws = wb.active
    assert ws.protection.sheet is True

    key_col = _find_col(ws, "Formula Code")
    editable_col = _find_col(ws, "Base Price")
    readonly_col = _find_col(ws, "Data Confidence")

    assert ws.cell(row=2, column=key_col).protection.locked is True
    assert ws.cell(row=2, column=editable_col).protection.locked is False
    assert ws.cell(row=2, column=readonly_col).protection.locked is True


def test_export_requires_formulas_edit_permission(client_as, tenant_a, coverage_setup):
    r = client_as(tenant_a).get("/api/sheets/formula_coverage_price/export",
                                 params={"subfamily_id": coverage_setup["sub_a"].id})
    assert r.status_code == 403


# ── AC2 ──────────────────────────────────────────────────────────────────

def test_reimport_unmodified_export_is_empty(client_as, admin, coverage_setup):
    c = client_as(admin)
    content = _export(c, subfamily_id=coverage_setup["sub_a"].id)
    r = _import(c, content, subfamily_id=coverage_setup["sub_a"].id)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["status"] == "empty"
    assert body["diffs"] == []


# ── AC3 ──────────────────────────────────────────────────────────────────

def test_reimport_edited_sheet_produces_named_diff_and_import_never_mutates(client_as, admin, coverage_setup, db):
    c = client_as(admin)
    content = _export(c, subfamily_id=coverage_setup["sub_a"].id)
    wb = _load(content)
    ws = wb.active
    price_col = _find_col(ws, "Base Price")
    code_col = _find_col(ws, "Formula Code")

    t1_code = coverage_setup["t1"].code
    for row in ws.iter_rows(min_row=2):
        if row[code_col - 1].value == t1_code:
            row[price_col - 1].value = 150

    r = _import(c, _save(wb), subfamily_id=coverage_setup["sub_a"].id)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["status"] == "diffed"
    changes = [d for d in body["diffs"] if d["kind"] == "change"]
    assert len(changes) == 1
    d = changes[0]
    assert d["row_key"] == {"code": t1_code, "region": "Europe"}
    assert d["column"] == "base_price"
    assert d["old_value"] == "100.0"
    assert d["new_value"] == "150.0"
    assert d["applied"] is False

    db.expire_all()
    cov = db.query(FormulaRegionCoverage).filter(FormulaRegionCoverage.template_id == coverage_setup["t1"].id).first()
    assert float(cov.base_price) == 100.0  # import never mutates

    run_id = body["id"]
    apply_r = c.post(f"/api/sheets/import-runs/{run_id}/apply")
    assert apply_r.status_code == 200, apply_r.text
    apply_body = apply_r.json()
    assert len(apply_body["applied"]) == 1
    assert apply_body["run"]["status"] == "applied"

    db.expire_all()
    cov = db.query(FormulaRegionCoverage).filter(FormulaRegionCoverage.template_id == coverage_setup["t1"].id).first()
    assert float(cov.base_price) == 150.0


# ── AC4 ──────────────────────────────────────────────────────────────────

def test_reordered_rows_still_rekey_by_business_key(client_as, admin, coverage_setup, db):
    c = client_as(admin)
    content = _export(c, subfamily_id=coverage_setup["sub_a"].id)
    wb = _load(content)
    ws = wb.active
    price_col = _find_col(ws, "Base Price")
    code_col = _find_col(ws, "Formula Code")

    data_rows = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
    assert len(data_rows) == 2
    data_rows.reverse()  # simulate a human sorting the sheet

    t2_code = coverage_setup["t2"].code
    for r_idx, row_values in enumerate(data_rows, start=2):
        for c_idx, val in enumerate(row_values, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
        if row_values[code_col - 1] == t2_code:
            ws.cell(row=r_idx, column=price_col, value=250)

    r = _import(c, _save(wb), subfamily_id=coverage_setup["sub_a"].id)
    assert r.status_code == 200, r.text
    changes = [d for d in r.json()["diffs"] if d["kind"] == "change"]
    assert len(changes) == 1
    assert changes[0]["row_key"]["code"] == t2_code
    assert changes[0]["new_value"] == "250.0"


# ── AC5 ──────────────────────────────────────────────────────────────────

def test_readonly_column_edit_is_rejected_not_applied(client_as, admin, coverage_setup, db):
    c = client_as(admin)
    content = _export(c, subfamily_id=coverage_setup["sub_a"].id)
    wb = _load(content)
    ws = wb.active
    conf_col = _find_col(ws, "Data Confidence")
    code_col = _find_col(ws, "Formula Code")
    t1_code = coverage_setup["t1"].code
    for row in ws.iter_rows(min_row=2):
        if row[code_col - 1].value == t1_code:
            row[conf_col - 1].value = "CONF-LOW"

    r = _import(c, _save(wb), subfamily_id=coverage_setup["sub_a"].id)
    assert r.status_code == 200, r.text
    body = r.json()
    rejected = [d for d in body["diffs"] if d["kind"] == "rejected_readonly_edit"]
    assert len(rejected) == 1
    assert rejected[0]["column"] == "data_confidence"
    assert not any(d["kind"] == "change" for d in body["diffs"])

    # Even calling apply must never touch it — only "change" diffs are appliable.
    c.post(f"/api/sheets/import-runs/{body['id']}/apply")
    db.expire_all()
    cov = db.query(FormulaRegionCoverage).filter(FormulaRegionCoverage.template_id == coverage_setup["t1"].id).first()
    assert cov.data_confidence == "CONF-HIGH"


# ── AC6 ──────────────────────────────────────────────────────────────────

def test_import_run_is_persisted_and_fetchable(client_as, admin, coverage_setup):
    c = client_as(admin)
    content = _export(c, subfamily_id=coverage_setup["sub_a"].id)
    r = _import(c, content, subfamily_id=coverage_setup["sub_a"].id)
    run_id = r.json()["id"]

    fresh = c.get(f"/api/sheets/import-runs/{run_id}")
    assert fresh.status_code == 200
    assert fresh.json()["id"] == run_id
    assert fresh.json()["status"] == "empty"


def test_list_import_runs_by_payload_key(client_as, admin, coverage_setup):
    c = client_as(admin)
    content = _export(c, subfamily_id=coverage_setup["sub_a"].id)
    _import(c, content, subfamily_id=coverage_setup["sub_a"].id)

    listed = c.get("/api/sheets/import-runs", params={"payload_key": "formula_coverage_price"})
    assert listed.status_code == 200
    assert len(listed.json()) >= 1


# ── Concurrency ──────────────────────────────────────────────────────────

def test_apply_skips_stale_row_when_live_value_changed_since_diff(client_as, admin, coverage_setup, db):
    c = client_as(admin)
    content = _export(c, subfamily_id=coverage_setup["sub_a"].id)
    wb = _load(content)
    ws = wb.active
    price_col = _find_col(ws, "Base Price")
    code_col = _find_col(ws, "Formula Code")
    t1_code = coverage_setup["t1"].code
    for row in ws.iter_rows(min_row=2):
        if row[code_col - 1].value == t1_code:
            row[price_col - 1].value = 150

    r = _import(c, _save(wb), subfamily_id=coverage_setup["sub_a"].id)
    run_id = r.json()["id"]

    # Simulate a second officer's concurrent change landing first.
    cov = db.query(FormulaRegionCoverage).filter(FormulaRegionCoverage.template_id == coverage_setup["t1"].id).first()
    cov.base_price = 999
    db.commit()

    apply_r = c.post(f"/api/sheets/import-runs/{run_id}/apply")
    assert apply_r.status_code == 200, apply_r.text
    body = apply_r.json()
    assert len(body["applied"]) == 0
    assert len(body["skipped_stale"]) == 1

    db.expire_all()
    cov = db.query(FormulaRegionCoverage).filter(FormulaRegionCoverage.template_id == coverage_setup["t1"].id).first()
    assert float(cov.base_price) == 999.0  # untouched by this apply


# ── Extras ───────────────────────────────────────────────────────────────

def test_invalid_value_reported_not_silently_dropped(client_as, admin, coverage_setup):
    c = client_as(admin)
    content = _export(c, subfamily_id=coverage_setup["sub_a"].id)
    wb = _load(content)
    ws = wb.active
    price_col = _find_col(ws, "Base Price")
    code_col = _find_col(ws, "Formula Code")
    t1_code = coverage_setup["t1"].code
    for row in ws.iter_rows(min_row=2):
        if row[code_col - 1].value == t1_code:
            row[price_col - 1].value = "not-a-number"

    r = _import(c, _save(wb), subfamily_id=coverage_setup["sub_a"].id)
    assert r.status_code == 200, r.text
    invalid = [d for d in r.json()["diffs"] if d["kind"] == "invalid_value"]
    assert len(invalid) == 1
    assert invalid[0]["column"] == "base_price"


def test_unmatched_key_reported(client_as, admin, coverage_setup):
    c = client_as(admin)
    content = _export(c, subfamily_id=coverage_setup["sub_a"].id)
    wb = _load(content)
    ws = wb.active
    code_col = _find_col(ws, "Formula Code")
    ws.cell(row=2, column=code_col, value="NO-SUCH-CODE")

    r = _import(c, _save(wb), subfamily_id=coverage_setup["sub_a"].id)
    assert r.status_code == 200, r.text
    unmatched = [d for d in r.json()["diffs"] if d["kind"] == "unmatched_key"]
    assert len(unmatched) == 1


def test_import_requires_formulas_edit_permission(client_as, tenant_a, coverage_setup):
    r = _import(client_as(tenant_a), b"irrelevant", subfamily_id=coverage_setup["sub_a"].id)
    assert r.status_code == 403


def test_apply_is_idempotent_on_already_applied_diffs(client_as, admin, coverage_setup, db):
    c = client_as(admin)
    content = _export(c, subfamily_id=coverage_setup["sub_a"].id)
    wb = _load(content)
    ws = wb.active
    price_col = _find_col(ws, "Base Price")
    code_col = _find_col(ws, "Formula Code")
    t1_code = coverage_setup["t1"].code
    for row in ws.iter_rows(min_row=2):
        if row[code_col - 1].value == t1_code:
            row[price_col - 1].value = 150

    r = _import(c, _save(wb), subfamily_id=coverage_setup["sub_a"].id)
    run_id = r.json()["id"]

    first = c.post(f"/api/sheets/import-runs/{run_id}/apply")
    assert len(first.json()["applied"]) == 1
    second = c.post(f"/api/sheets/import-runs/{run_id}/apply")
    assert second.status_code == 200
    assert len(second.json()["applied"]) == 0
    assert len(second.json()["skipped_stale"]) == 0

    db.expire_all()
    cov = db.query(FormulaRegionCoverage).filter(FormulaRegionCoverage.template_id == coverage_setup["t1"].id).first()
    assert float(cov.base_price) == 150.0
